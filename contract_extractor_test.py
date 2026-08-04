"""
Contract Extraction - Test/Demo Mode
Tests the extraction pipeline without API calls using mock data
"""

import json
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pypdf import PdfReader
from docx import Document
from pydantic import BaseModel, Field


class ContractHeader(BaseModel):
    """Schema for contract header information"""
    contract_number: str = Field(default="", description="Contract reference number")
    vendor_name: str = Field(default="", description="Vendor/Supplier name")
    start_date: str = Field(default="", description="Contract start date")
    end_date: str = Field(default="", description="Contract end date")
    contract_value: str = Field(default="", description="Total contract value")
    payment_terms: str = Field(default="", description="Payment terms (e.g., Net 30)")
    currency: str = Field(default="", description="Currency of contract")
    contract_type: str = Field(default="", description="Type of contract")


class ServiceDetail(BaseModel):
    """Schema for service/rate information"""
    contract_number: str = Field(default="", description="Contract reference number")
    service_name: str = Field(default="", description="Name of service")
    service_description: str = Field(default="", description="Description of service")
    unit: str = Field(default="", description="Unit of measurement (e.g., per hour, per unit)")
    rate: str = Field(default="", description="Cost/rate for the service")
    currency: str = Field(default="", description="Currency of rate")
    minimum_order: str = Field(default="", description="Minimum order quantity/value")
    volume_discount: str = Field(default="", description="Volume discount terms if applicable")
    effective_from: str = Field(default="", description="Date when rate is effective")


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file"""
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        print(f"Error reading PDF {file_path}: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word document"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error reading DOCX {file_path}: {e}")
        return ""


def load_contract_text(file_path: str) -> str:
    """Load contract text from PDF or Word document"""
    file_path = str(file_path)

    if file_path.lower().endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_path.lower().endswith('.docx'):
        return extract_text_from_docx(file_path)
    else:
        print(f"Unsupported file format: {file_path}")
        return ""


# Mock extraction data for demo purposes
MOCK_EXTRACTIONS = {
    "sample_contract_1_techsolutions.docx": {
        "contract_header": {
            "contract_number": "VC-2025-0001",
            "vendor_name": "TechSolutions Inc.",
            "start_date": "January 1, 2025",
            "end_date": "December 31, 2025",
            "contract_value": "$150,000 USD",
            "payment_terms": "Net 30 days",
            "currency": "USD",
            "contract_type": "Service Agreement"
        },
        "services": [
            {
                "service_name": "Cloud Infrastructure Support",
                "service_description": "24/7 monitoring, maintenance, and technical support",
                "unit": "Per Month",
                "rate": "$5,000",
                "currency": "USD",
                "minimum_order": "1 month",
                "volume_discount": "10% for annual commitment",
                "effective_from": "January 1, 2025"
            },
            {
                "service_name": "Security Audit & Penetration Testing",
                "service_description": "Quarterly security audits and vulnerability assessments",
                "unit": "Per Quarter",
                "rate": "$8,000",
                "currency": "USD",
                "minimum_order": "1 quarter",
                "volume_discount": "15% for 4-quarter commitment",
                "effective_from": "January 1, 2025"
            },
            {
                "service_name": "On-Site Technical Support",
                "service_description": "On-site technical support for critical infrastructure",
                "unit": "Per Hour",
                "rate": "$150",
                "currency": "USD",
                "minimum_order": "4 hours",
                "volume_discount": "10% for 100+ hours",
                "effective_from": "January 1, 2025"
            }
        ]
    },
    "sample_contract_2_officeworld.docx": {
        "contract_header": {
            "contract_number": "VC-2025-0002",
            "vendor_name": "OfficeWorld Supplies Ltd.",
            "start_date": "February 1, 2025",
            "end_date": "January 31, 2026",
            "contract_value": "$45,000 USD",
            "payment_terms": "Net 15 days",
            "currency": "USD",
            "contract_type": "Supply Agreement"
        },
        "services": [
            {
                "service_name": "Premium Printer Paper A4 80gsm",
                "service_description": "High quality white printer paper, A4 size",
                "unit": "Per Ream",
                "rate": "$5.50",
                "currency": "USD",
                "minimum_order": "10 reams",
                "volume_discount": "5% for 100+ reams, 10% for 500+",
                "effective_from": "February 1, 2025"
            },
            {
                "service_name": "Ballpoint Pens - Black/Blue/Red",
                "service_description": "Professional ballpoint pens in multiple colors",
                "unit": "Per Box",
                "rate": "$12.00",
                "currency": "USD",
                "minimum_order": "5 boxes",
                "volume_discount": "8% for monthly orders of 200+ boxes",
                "effective_from": "February 1, 2025"
            }
        ]
    },
    "sample_contract_3_globalship.docx": {
        "contract_header": {
            "contract_number": "VC-2025-0003",
            "vendor_name": "GlobalShip Logistics",
            "start_date": "March 15, 2025",
            "end_date": "March 14, 2026",
            "contract_value": "$250,000 USD",
            "payment_terms": "Net 45 days",
            "currency": "USD",
            "contract_type": "Logistics Services Agreement"
        },
        "services": [
            {
                "service_name": "Domestic Ground Shipping",
                "service_description": "Standard ground shipping within continental US",
                "unit": "Per Pound",
                "rate": "$0.85",
                "currency": "USD",
                "minimum_order": "100 pounds",
                "volume_discount": "10% for monthly volumes over 10,000 lbs",
                "effective_from": "March 15, 2025"
            },
            {
                "service_name": "International Express Shipping",
                "service_description": "Express international shipping to multiple regions",
                "unit": "Per Pound",
                "rate": "$2.50",
                "currency": "USD",
                "minimum_order": "50 pounds",
                "volume_discount": "15% for monthly international volumes over 5,000 lbs",
                "effective_from": "March 15, 2025"
            },
            {
                "service_name": "Warehouse Storage & Fulfillment",
                "service_description": "Climate-controlled warehouse storage with fulfillment",
                "unit": "Per Pallet per Month",
                "rate": "$75",
                "currency": "USD",
                "minimum_order": "10 pallets",
                "volume_discount": "20% for 6-month commitment of 50+ pallets",
                "effective_from": "March 15, 2025"
            }
        ]
    }
}


def extract_contract_info_mock(file_name: str) -> tuple[Optional[ContractHeader], list[ServiceDetail]]:
    """
    Mock extraction - returns predefined data for demo/testing
    In production, this would call Claude AI
    """
    print(f"  [TEST MODE] Using mock data for {file_name}")

    # Use specific mock data if available, otherwise use default
    if file_name in MOCK_EXTRACTIONS:
        data = MOCK_EXTRACTIONS[file_name]
    else:
        # Default mock data for any unrecognized filename
        print(f"  [INFO] Using default mock data for {file_name}")
        data = {
            "contract_header": {
                "contract_number": f"VC-{file_name[:10].upper()}",
                "vendor_name": "Vendor Name",
                "start_date": "January 1, 2025",
                "end_date": "December 31, 2025",
                "contract_value": "$200,000 USD",
                "payment_terms": "Net 30 days",
                "currency": "USD",
                "contract_type": "Service Agreement"
            },
            "services": [
                {
                    "service_name": "Professional Services",
                    "service_description": "Consulting and support services",
                    "unit": "Per Hour",
                    "rate": "$150",
                    "currency": "USD",
                    "minimum_order": "40 hours",
                    "volume_discount": "5% for orders over 100 hours",
                    "effective_from": "January 1, 2025"
                },
                {
                    "service_name": "Technical Support",
                    "service_description": "24/7 technical support",
                    "unit": "Per Month",
                    "rate": "$5,000",
                    "currency": "USD",
                    "minimum_order": "1 month",
                    "volume_discount": "10% for 6+ month commitments",
                    "effective_from": "January 1, 2025"
                }
            ]
        }

    # Create ContractHeader object
    header_data = data.get("contract_header", {})
    contract_header = ContractHeader(**header_data)

    # Create ServiceDetail objects
    services = []
    for service_data in data.get("services", []):
        service_data["contract_number"] = contract_header.contract_number
        services.append(ServiceDetail(**service_data))

    return contract_header, services


def create_excel_output(all_headers: list[ContractHeader], all_services: list[ServiceDetail], output_file: str = "extracted_contracts.xlsx"):
    """
    Create Excel workbook with 2 sheets:
    Sheet 1: Contract Headers
    Sheet 2: Services & Rates
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Contract Headers
    ws_headers = wb.create_sheet("Contract Headers")
    header_cols = ["Contract Number", "Vendor Name", "Start Date", "End Date",
                   "Contract Value", "Payment Terms", "Currency", "Contract Type"]
    ws_headers.append(header_cols)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_headers[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for header in all_headers:
        if header.contract_number:
            ws_headers.append([
                header.contract_number,
                header.vendor_name,
                header.start_date,
                header.end_date,
                header.contract_value,
                header.payment_terms,
                header.currency,
                header.contract_type
            ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_headers.column_dimensions[col].width = 18

    # Sheet 2: Services & Rates
    ws_services = wb.create_sheet("Services & Rates")
    service_cols = ["Contract Number", "Service Name", "Description", "Unit",
                    "Rate", "Currency", "Minimum Order", "Volume Discount", "Effective From"]
    ws_services.append(service_cols)

    for cell in ws_services[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for service in all_services:
        ws_services.append([
            service.contract_number,
            service.service_name,
            service.service_description,
            service.unit,
            service.rate,
            service.currency,
            service.minimum_order,
            service.volume_discount,
            service.effective_from
        ])

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_services.column_dimensions[col].width = 18

    wb.save(output_file)
    print(f"[OK] Excel file created: {output_file}")


def process_contracts_test(input_folder: str = "sample-contracts", output_file: str = "extracted_contracts.xlsx"):
    """
    Test mode: Process all contracts using mock data
    """
    input_path = Path(input_folder)

    if not input_path.exists():
        print(f"Input folder not found: {input_folder}")
        return

    contract_files = list(input_path.glob("*.pdf")) + list(input_path.glob("*.docx"))

    if not contract_files:
        print(f"No PDF or DOCX files found in {input_folder}")
        return

    print(f"Found {len(contract_files)} contract(s) to process")

    all_headers = []
    all_services = []

    for idx, file_path in enumerate(contract_files, 1):
        print(f"\n[{idx}/{len(contract_files)}] Processing: {file_path.name}")

        # Extract info using mock data
        header, services = extract_contract_info_mock(file_path.name)

        if header and header.contract_number:
            all_headers.append(header)
            all_services.extend(services)
            print(f"  [OK] Extracted: {header.contract_number} - {header.vendor_name}")
            print(f"    Services found: {len(services)}")
        else:
            print(f"  [WARN] No contract information extracted")

    if all_headers or all_services:
        create_excel_output(all_headers, all_services, output_file)
        print(f"\n[OK] Test complete!")
        print(f"  Total contracts: {len(all_headers)}")
        print(f"  Total services: {len(all_services)}")
        print(f"\n[OUTPUT] Output file: {output_file}")
        print(f"\nWhen you have an API key, use contract_extractor.py instead for real extraction.")
    else:
        print("\n[WARN] No contract data extracted.")


if __name__ == "__main__":
    print("=" * 60)
    print("CONTRACT EXTRACTION - TEST MODE (No API calls)")
    print("=" * 60)
    print()

    process_contracts_test(
        input_folder="sample-contracts",
        output_file="extracted_contracts_test.xlsx"
    )
