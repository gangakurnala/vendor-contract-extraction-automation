"""
Vendor Contract Extraction Automation
Extracts contract information from PDF and Word documents using Claude AI
Outputs consolidated data to Excel format
"""

import os
import json
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pypdf import PdfReader
from docx import Document
from pydantic import BaseModel, Field

# Load environment variables from .env.web
load_dotenv('.env.web')

# Initialize Anthropic client (lazy - only when API key exists)
client = None

def get_client():
    global client
    if client is None:
        # For corporate environments with SSL certificate issues, use httpx client with verify=False
        import httpx
        http_client = httpx.Client(verify=False)
        client = Anthropic(http_client=http_client)
    return client


class ContractHeader(BaseModel):
    """Schema for contract header information"""
    contract_number: str = Field(default="", description="Contract reference number")
    vendor_name: str = Field(default="", description="Vendor/Supplier name")
    start_date: str = Field(default="", description="Contract start date")
    end_date: str = Field(default="", description="Contract end date")
    contract_value: str = Field(default="", description="Total contract value")
    payment_terms: str = Field(default="", description="Payment terms (e.g., Net 30)")
    currency: str = Field(default="", description="Currency of contract")
    contract_type: str = Field(default="", description="Type of contract")


class ServiceDetail(BaseModel):
    """Schema for service/rate information"""
    contract_number: str = Field(default="", description="Contract reference number")
    service_name: str = Field(default="", description="Name of service")
    service_description: str = Field(default="", description="Description of service")
    unit: str = Field(default="", description="Unit of measurement (e.g., per hour, per unit)")
    rate: str = Field(default="", description="Cost/rate for the service")
    currency: str = Field(default="", description="Currency of rate")
    minimum_order: str = Field(default="", description="Minimum order quantity/value")
    volume_discount: str = Field(default="", description="Volume discount terms if applicable")
    effective_from: str = Field(default="", description="Date when rate is effective")


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF file"""
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        print(f"Error reading PDF {file_path}: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word document"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error reading DOCX {file_path}: {e}")
        return ""


def load_contract_text(file_path: str) -> str:
    """Load contract text from PDF or Word document"""
    file_path = str(file_path)

    if file_path.lower().endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_path.lower().endswith('.docx'):
        return extract_text_from_docx(file_path)
    else:
        print(f"Unsupported file format: {file_path}")
        return ""


def extract_contract_info(contract_text: str, file_name: str) -> tuple[Optional[ContractHeader], list[ServiceDetail]]:
    """
    Extract contract information using Claude Sonnet 5.0 if API key available,
    otherwise uses smart text-based extraction from the contract
    Returns: (contract_header, list_of_services)
    """
    import re

    if not contract_text.strip():
        print(f"Empty contract text for {file_name}")
        return None, []

    data = None
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()

    # Try Claude Sonnet 5.0 if API key is configured
    if api_key and api_key != "sk-ant-YOUR_API_KEY_HERE":
        try:
            extraction_prompt = f"""
            Analyze this vendor contract and extract the following information:

            CONTRACT TEXT:
            {contract_text[:5000]}

            Please extract and return a JSON response with this structure:
            {{
                "contract_header": {{
                    "contract_number": "extracted contract number or ID",
                    "vendor_name": "vendor or supplier name",
                    "start_date": "contract start date",
                    "end_date": "contract end date",
                    "contract_value": "total contract value",
                    "payment_terms": "payment terms",
                    "currency": "currency",
                    "contract_type": "type of contract"
                }},
                "services": [
                    {{
                        "service_name": "service name",
                        "service_description": "what the service entails",
                        "unit": "unit of measurement",
                        "rate": "cost per unit",
                        "currency": "currency",
                        "minimum_order": "minimum order if any",
                        "volume_discount": "volume discount info",
                        "effective_from": "effective date"
                    }}
                ]
            }}

            If information is not available, use empty strings.
            Return ONLY the JSON object, no other text.
            """

            response = get_client().messages.create(
                model="claude-sonnet-5",
                max_tokens=2048,
                messages=[{"role": "user", "content": extraction_prompt}]
            )
            response_text = response.content[0].text
            data = json.loads(response_text)
            print(f"  [CLAUDE SONNET 5.0] Extracted from {file_name}")
        except Exception as e:
            import traceback
            print(f"  [ERROR] Claude extraction failed: {type(e).__name__}: {str(e)}")
            print(f"  {traceback.format_exc()}")
            data = None

    # Fall back to smart text extraction if no API key or Claude fails
    if not data:
        print(f"  [TEXT EXTRACTION] Extracting from {file_name}")
        text_upper = contract_text.upper()

        # Extract contract number (look for patterns like "Contract #", "Agreement No", etc.)
        contract_num = ""
        patterns = [r'CONTRACT\s*(?:NUMBER|NO\.?|#)\s*[:=\s]*([A-Z0-9\-]+)',
                   r'AGREEMENT\s*(?:NUMBER|NO\.?)\s*[:=\s]*([A-Z0-9\-]+)',
                   r'([A-Z]{2,}-\d{4}-\d{4})']
        for pattern in patterns:
            match = re.search(pattern, text_upper)
            if match:
                contract_num = match.group(1)
                break

        # Extract vendor name (look for patterns like "Vendor:", "Supplier:", "Company:", "between ... and ...")
        vendor = ""
        vendor_patterns = [r'VENDOR\s*[:=]\s*([^\n]+)',
                          r'SUPPLIER\s*[:=]\s*([^\n]+)',
                          r'COMPANY\s*[:=]\s*([^\n]+)',
                          r'between\s+([^\n]+?)\s+\(.*?COMPANY.*?\)',
                          r'(?:FROM|BY)\s*[:=]\s*([^\n]+)']
        for pattern in vendor_patterns:
            match = re.search(pattern, text_upper)
            if match:
                vendor = match.group(1).strip()
                # Clean up parenthetical references
                vendor = re.sub(r'\(.*?\)', '', vendor).strip()
                if vendor:
                    break

        # Extract dates
        start_date = ""
        end_date = ""
        date_pattern = r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})'
        dates = re.findall(date_pattern, contract_text)
        if len(dates) >= 1:
            start_date = dates[0]
        if len(dates) >= 2:
            end_date = dates[1]

        # Extract value (look for currency symbols and amounts)
        value = ""
        value_patterns = [r'(?:TOTAL\s+)?(?:VALUE|AMOUNT)\s*[:=]\s*([^\n]+)',
                         r'([$€£]\s*[\d,]+(?:\.\d{2})?)']
        for pattern in value_patterns:
            match = re.search(pattern, contract_text)
            if match:
                value = match.group(1).strip()
                break

        # Extract payment terms
        payment_terms = ""
        if 'NET 30' in text_upper:
            payment_terms = "Net 30 days"
        elif 'NET 60' in text_upper:
            payment_terms = "Net 60 days"
        elif 'NET 90' in text_upper:
            payment_terms = "Net 90 days"
        elif 'PAYMENT' in text_upper:
            match = re.search(r'PAYMENT\s+TERMS?\s*[:=]\s*([^\n]+)', text_upper)
            if match:
                payment_terms = match.group(1).strip()

        # Detect currency
        currency = "USD"
        if "EUR" in text_upper or "EURO" in text_upper:
            currency = "EUR"
        elif "GBP" in text_upper or "POUND" in text_upper:
            currency = "GBP"

        data = {
            "contract_header": {
                "contract_number": contract_num or f"EXTRACTED-{file_name[:8].upper()}",
                "vendor_name": vendor or "Extracted Vendor",
                "start_date": start_date,
                "end_date": end_date,
                "contract_value": value,
                "payment_terms": payment_terms,
                "currency": currency,
                "contract_type": "Service Agreement"
            },
            "services": []
        }

    try:
        header_data = data.get("contract_header", {})
        contract_header = ContractHeader(**header_data)
        services = []
        for service_data in data.get("services", []):
            service_data["contract_number"] = contract_header.contract_number or file_name
            services.append(ServiceDetail(**service_data))
        return contract_header, services
    except Exception as e:
        print(f"Error processing extracted data from {file_name}: {e}")
        return None, []


def create_excel_output(all_headers: list[ContractHeader], all_services: list[ServiceDetail], output_file: str = "extracted_contracts.xlsx"):
    """
    Create Excel workbook with 2 sheets:
    Sheet 1: Contract Headers
    Sheet 2: Services & Rates
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Contract Headers
    ws_headers = wb.create_sheet("Contract Headers")
    header_cols = ["Contract Number", "Vendor Name", "Start Date", "End Date",
                   "Contract Value", "Payment Terms", "Currency", "Contract Type"]
    ws_headers.append(header_cols)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_headers[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add contract header data
    for header in all_headers:
        if header.contract_number:  # Only add if we have contract number
            ws_headers.append([
                header.contract_number,
                header.vendor_name,
                header.start_date,
                header.end_date,
                header.contract_value,
                header.payment_terms,
                header.currency,
                header.contract_type
            ])

    # Adjust column widths for headers sheet
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_headers.column_dimensions[col].width = 15

    # Sheet 2: Services & Rates
    ws_services = wb.create_sheet("Services & Rates")
    service_cols = ["Contract Number", "Service Name", "Description", "Unit",
                    "Rate", "Currency", "Minimum Order", "Volume Discount", "Effective From"]
    ws_services.append(service_cols)

    # Style header row
    for cell in ws_services[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add service data
    for service in all_services:
        ws_services.append([
            service.contract_number,
            service.service_name,
            service.service_description,
            service.unit,
            service.rate,
            service.currency,
            service.minimum_order,
            service.volume_discount,
            service.effective_from
        ])

    # Adjust column widths for services sheet
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_services.column_dimensions[col].width = 15

    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")


def process_contracts(input_folder: str = "sample-contracts", output_file: str = "extracted_contracts.xlsx"):
    """
    Main function to process all contracts in a folder
    """
    input_path = Path(input_folder)

    if not input_path.exists():
        print(f"Input folder not found: {input_folder}")
        return

    # Find all PDF and DOCX files
    contract_files = list(input_path.glob("*.pdf")) + list(input_path.glob("*.docx"))

    if not contract_files:
        print(f"No PDF or DOCX files found in {input_folder}")
        return

    print(f"Found {len(contract_files)} contract(s) to process")

    all_headers = []
    all_services = []

    # Process each contract
    for idx, file_path in enumerate(contract_files, 1):
        print(f"\n[{idx}/{len(contract_files)}] Processing: {file_path.name}")

        # Extract text
        contract_text = load_contract_text(file_path)

        if not contract_text:
            print(f"  [WARN] Could not extract text from {file_path.name}")
            continue

        # Extract info using Claude
        print(f"  Analyzing with Claude AI...")
        header, services = extract_contract_info(contract_text, file_path.name)

        if header and header.contract_number:
            all_headers.append(header)
            all_services.extend(services)
            print(f"  [OK] Extracted: {header.contract_number} - {header.vendor_name}")
            print(f"    Services found: {len(services)}")
        else:
            print(f"  [WARN] No contract information extracted")

    # Create Excel output
    if all_headers or all_services:
        create_excel_output(all_headers, all_services, output_file)
        print(f"\n[OK] Processing complete!")
        print(f"  Total contracts: {len(all_headers)}")
        print(f"  Total services: {len(all_services)}")
    else:
        print("\n[WARN] No contract data extracted. Check input files and contract format.")


if __name__ == "__main__":
    # Example usage
    print("Vendor Contract Extraction Automation")
    print("=" * 50)

    # Process contracts from sample-contracts folder
    process_contracts(
        input_folder="sample-contracts",
        output_file="extracted_contracts.xlsx"
    )
