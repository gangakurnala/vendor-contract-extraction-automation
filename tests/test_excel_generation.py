"""
Unit tests for Excel generation functionality
Tests workbook creation, sheet formatting, and data output
"""

import pytest
from pathlib import Path
import sys
import tempfile
import os

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from contract_extractor import (
    create_excel_output,
    ContractHeader,
    ServiceDetail
)
from openpyxl import load_workbook


class TestExcelGeneration:
    """Test Excel file generation"""

    @pytest.fixture
    def sample_headers(self):
        """Create sample contract headers for testing"""
        return [
            ContractHeader(
                contract_number="VC-2025-0001",
                vendor_name="TechSolutions Inc.",
                start_date="2025-01-01",
                end_date="2025-12-31",
                contract_value="$150,000",
                payment_terms="Net 30",
                currency="USD",
                contract_type="Service Agreement"
            ),
            ContractHeader(
                contract_number="VC-2025-0002",
                vendor_name="OfficeWorld Supplies",
                start_date="2025-02-01",
                end_date="2026-01-31",
                contract_value="$45,000",
                payment_terms="Net 15",
                currency="USD",
                contract_type="Supply Agreement"
            )
        ]

    @pytest.fixture
    def sample_services(self):
        """Create sample service details for testing"""
        return [
            ServiceDetail(
                contract_number="VC-2025-0001",
                service_name="Cloud Support",
                service_description="24/7 cloud support",
                unit="Per Month",
                rate="$5,000",
                currency="USD",
                minimum_order="1 month",
                volume_discount="10%",
                effective_from="2025-01-01"
            ),
            ServiceDetail(
                contract_number="VC-2025-0001",
                service_name="Security Audit",
                service_description="Quarterly audit",
                unit="Per Quarter",
                rate="$8,000",
                currency="USD",
                minimum_order="1 quarter",
                volume_discount="15%",
                effective_from="2025-01-01"
            ),
            ServiceDetail(
                contract_number="VC-2025-0002",
                service_name="Paper Supplies",
                service_description="Printer paper",
                unit="Per Ream",
                rate="$5.50",
                currency="USD",
                minimum_order="10 reams",
                volume_discount="5%",
                effective_from="2025-02-01"
            )
        ]

    def test_excel_creation_basic(self, sample_headers, sample_services):
        """Test basic Excel file creation"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_output.xlsx")

            # Generate Excel
            create_excel_output(sample_headers, sample_services, output_file)

            # Verify file exists
            assert os.path.exists(output_file)
            assert os.path.getsize(output_file) > 0

    def test_excel_has_two_sheets(self, sample_headers, sample_services):
        """Test that Excel has 2 sheets"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_output.xlsx")

            create_excel_output(sample_headers, sample_services, output_file)

            # Load and verify sheets
            wb = load_workbook(output_file)
            assert len(wb.sheetnames) == 2
            assert "Contract Headers" in wb.sheetnames
            assert "Services & Rates" in wb.sheetnames

    def test_excel_headers_sheet_content(self, sample_headers, sample_services):
        """Test Contract Headers sheet content"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_output.xlsx")

            create_excel_output(sample_headers, sample_services, output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]

            # Check header row
            assert ws_headers["A1"].value == "Contract Number"
            assert ws_headers["B1"].value == "Vendor Name"
            assert ws_headers["C1"].value == "Start Date"

            # Check data rows
            assert ws_headers["A2"].value == "VC-2025-0001"
            assert ws_headers["B2"].value == "TechSolutions Inc."
            assert ws_headers["A3"].value == "VC-2025-0002"
            assert ws_headers["B3"].value == "OfficeWorld Supplies"

    def test_excel_services_sheet_content(self, sample_headers, sample_services):
        """Test Services & Rates sheet content"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_output.xlsx")

            create_excel_output(sample_headers, sample_services, output_file)

            wb = load_workbook(output_file)
            ws_services = wb["Services & Rates"]

            # Check header row
            assert ws_services["A1"].value == "Contract Number"
            assert ws_services["B1"].value == "Service Name"
            assert ws_services["C1"].value == "Description"
            assert ws_services["E1"].value == "Rate"

            # Check data rows
            assert ws_services["B2"].value == "Cloud Support"
            assert ws_services["E2"].value == "$5,000"
            assert ws_services["B3"].value == "Security Audit"

    def test_excel_empty_data(self):
        """Test Excel generation with empty data"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_empty.xlsx")

            create_excel_output([], [], output_file)

            # File should still be created
            assert os.path.exists(output_file)

            wb = load_workbook(output_file)
            assert len(wb.sheetnames) == 2

    def test_excel_only_headers(self, sample_headers):
        """Test Excel with only contract headers, no services"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_headers_only.xlsx")

            create_excel_output(sample_headers, [], output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]

            # Should have header + data rows
            assert ws_headers["A2"].value == "VC-2025-0001"
            assert ws_headers["A3"].value == "VC-2025-0002"

    def test_excel_only_services(self, sample_services):
        """Test Excel with only services, no headers"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_services_only.xlsx")

            create_excel_output([], sample_services, output_file)

            wb = load_workbook(output_file)
            ws_services = wb["Services & Rates"]

            # Should have service data
            assert ws_services["B2"].value == "Cloud Support"

    def test_excel_column_widths(self, sample_headers, sample_services):
        """Test that columns have appropriate widths"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_widths.xlsx")

            create_excel_output(sample_headers, sample_services, output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]

            # Check column widths are set
            assert ws_headers.column_dimensions["A"].width == 18
            assert ws_headers.column_dimensions["B"].width == 18

    def test_excel_header_formatting(self, sample_headers, sample_services):
        """Test that header rows have formatting"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_formatting.xlsx")

            create_excel_output(sample_headers, sample_services, output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]

            # Check header row formatting
            cell_a1 = ws_headers["A1"]
            assert cell_a1.font.bold is True
            assert cell_a1.font.color.rgb == "FFFFFFFF"  # White
            assert cell_a1.fill.start_color.rgb == "FF4472C4"  # Blue

    def test_excel_multiple_services_per_contract(self):
        """Test Excel with multiple services per contract"""
        headers = [
            ContractHeader(
                contract_number="VC-2025-0001",
                vendor_name="Vendor 1"
            )
        ]

        services = [
            ServiceDetail(
                contract_number="VC-2025-0001",
                service_name=f"Service {i}",
                rate=f"${1000 * i}"
            ) for i in range(1, 6)
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_multi_service.xlsx")

            create_excel_output(headers, services, output_file)

            wb = load_workbook(output_file)
            ws_services = wb["Services & Rates"]

            # Should have 5 services
            assert ws_services["B2"].value == "Service 1"
            assert ws_services["B6"].value == "Service 5"
            assert ws_services["E2"].value == "$1000"
            assert ws_services["E6"].value == "$5000"

    def test_excel_large_dataset(self):
        """Test Excel generation with larger dataset"""
        headers = [
            ContractHeader(
                contract_number=f"VC-2025-{i:04d}",
                vendor_name=f"Vendor {i}",
                contract_value=f"${100000 * i}"
            ) for i in range(1, 21)
        ]

        services = [
            ServiceDetail(
                contract_number=f"VC-2025-{i:04d}",
                service_name=f"Service {j}",
                rate=f"${5000 * j}"
            ) for i in range(1, 21) for j in range(1, 4)
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_large.xlsx")

            create_excel_output(headers, services, output_file)

            # Verify file was created
            assert os.path.exists(output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]
            ws_services = wb["Services & Rates"]

            # Check data
            assert ws_headers["A2"].value == "VC-2025-0001"
            assert ws_headers["A21"].value == "VC-2025-0020"
            assert ws_services["B2"].value == "Service 1"

    def test_excel_special_characters(self):
        """Test Excel with special characters in data"""
        headers = [
            ContractHeader(
                contract_number="VC-2025-0001",
                vendor_name="ABC & XYZ Inc.",
                contract_value="$100,000.50"
            )
        ]

        services = [
            ServiceDetail(
                contract_number="VC-2025-0001",
                service_name="Support & Maintenance",
                service_description="24/7 support & updates",
                rate="$5,000/month"
            )
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_special_chars.xlsx")

            create_excel_output(headers, services, output_file)

            wb = load_workbook(output_file)
            ws_headers = wb["Contract Headers"]

            # Check special characters preserved
            assert ws_headers["B2"].value == "ABC & XYZ Inc."

    def test_excel_unicode_characters(self):
        """Test Excel with unicode characters"""
        headers = [
            ContractHeader(
                contract_number="VC-2025-0001",
                vendor_name="Tech Solutions™ 中文"
            )
        ]

        services = [
            ServiceDetail(
                contract_number="VC-2025-0001",
                service_name="云服务 (Cloud Services) – Premium",
                rate="¥5,000"
            )
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test_unicode.xlsx")

            # Should handle unicode gracefully
            try:
                create_excel_output(headers, services, output_file)
                assert os.path.exists(output_file)
            except Exception as e:
                pytest.fail(f"Unicode handling failed: {e}")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
