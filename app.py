"""
Flask Application - Frontend Web UI Only
Contract Extraction Web Application
"""

from flask import Flask, render_template, request, jsonify, send_file, current_app, session, redirect
from flask_cors import CORS
from config import get_config
from models import db, User, ExtractionJob, JobStatus, ContractResult, AuditLog
from auth import get_auth_provider, create_tokens, log_audit
from contract_extractor import process_contracts
from werkzeug.utils import secure_filename
import os
import uuid
from datetime import datetime, timedelta
import logging
from pathlib import Path

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def create_app(config_name='development'):
    """Application factory"""

    app = Flask(__name__)
    app.config.from_object(get_config())
    app.secret_key = app.config.get('SECRET_KEY', 'dev-secret')
    app.permanent_session_lifetime = timedelta(days=7)

    # Initialize extensions
    CORS(app)
    if not app.extensions.get('sqlalchemy'):
        db.init_app(app)

    # Initialize database
    with app.app_context():
        db.create_all()

    # ==================== Helper Functions ====================

    def allowed_file(filename):
        """Check if file extension is allowed"""
        allowed_extensions = {'pdf', 'docx'}
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

    def get_current_user():
        """Get current user from session"""
        if 'user_id' not in session:
            return None
        return User.query.get(session.get('user_id'))

    # ==================== Web Routes ====================

    @app.route('/')
    def index():
        """Home/Login page"""
        if 'user_id' in session:
            return redirect('/dashboard')
        return render_template('index.html')

    @app.route('/login', methods=['POST'])
    def login():
        """Handle login"""
        try:
            username = request.form.get('username')
            password = request.form.get('password')

            if not username or not password:
                return jsonify({'error': 'Missing username or password'}), 400

            # Authenticate user
            auth_provider = get_auth_provider()
            user_info, status_code = auth_provider.authenticate(username, password)

            if status_code != 200:
                return jsonify(user_info), status_code

            # Create session
            user_id = user_info['user_id']
            session['user_id'] = user_id
            session['username'] = user_info['username']
            session.permanent = True

            # Log audit
            log_audit(user_id, 'LOGIN', 'user', user_id)

            return jsonify({'message': 'Login successful', 'user': user_info}), 200

        except Exception as e:
            logger.error(f"Login error: {e}")
            return jsonify({'error': 'Login failed'}), 500

    @app.route('/logout')
    def logout():
        """Handle logout"""
        if 'user_id' in session:
            log_audit(session['user_id'], 'LOGOUT', 'user', session['user_id'])
        session.clear()
        return redirect('/')

    @app.route('/dashboard')
    def dashboard():
        """User dashboard"""
        if 'user_id' not in session:
            return redirect('/')
        return render_template('dashboard.html')

    @app.route('/upload')
    def upload_page():
        """Upload page"""
        if 'user_id' not in session:
            return redirect('/')
        return render_template('upload.html')

    @app.route('/jobs')
    def jobs_page():
        """Jobs list page"""
        if 'user_id' not in session:
            return redirect('/')
        return render_template('jobs.html')

    # ==================== AJAX Endpoints (Internal Only) ====================

    @app.route('/api/user')
    def get_user():
        """Get current user info"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        user = User.query.get(session['user_id'])
        if user:
            return jsonify(user.to_dict()), 200
        return jsonify({'error': 'User not found'}), 404

    @app.route('/api/extraction/upload', methods=['POST'])
    def upload_files():
        """Upload contract files"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']

            # Check if files provided
            if 'files' not in request.files:
                return jsonify({'error': 'No files provided'}), 400

            files = request.files.getlist('files')
            if not files or files[0].filename == '':
                return jsonify({'error': 'No files selected'}), 400

            # Create job
            job_id = str(uuid.uuid4())
            job_name = request.form.get('job_name', f'Extraction {job_id[:8]}')

            job = ExtractionJob(
                id=job_id,
                user_id=user_id,
                job_name=job_name,
                description=request.form.get('description', ''),
                status=JobStatus.PENDING.value,
                input_file_count=len(files)
            )

            # Save uploaded files
            job_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], job_id)
            os.makedirs(job_folder, exist_ok=True)

            uploaded_files = []
            for file in files:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(job_folder, filename)
                    file.save(filepath)
                    uploaded_files.append(filename)

            if not uploaded_files:
                return jsonify({'error': 'No valid files uploaded'}), 400

            job.uploaded_files = uploaded_files
            db.session.add(job)
            db.session.commit()

            # Log audit
            log_audit(user_id, 'UPLOAD', 'job', job_id, {'file_count': len(uploaded_files)})

            return jsonify({
                'message': 'Files uploaded successfully',
                'job_id': job_id,
                'job': job.to_dict()
            }), 201

        except Exception as e:
            logger.error(f"Upload error: {e}")
            return jsonify({'error': 'Upload failed'}), 500

    @app.route('/api/extraction/extract/<job_id>', methods=['POST'])
    def extract_contracts(job_id):
        """Start extraction process"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']

            # Get job
            job = ExtractionJob.query.get(job_id)
            if not job:
                return jsonify({'error': 'Job not found'}), 404

            if job.user_id != user_id:
                return jsonify({'error': 'Unauthorized'}), 403

            # Start extraction
            job.status = JobStatus.PROCESSING.value
            job.started_at = datetime.utcnow()
            db.session.commit()

            # Process files with text extraction (or Claude if API key available)
            job_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], job_id)
            output_file = os.path.join(
                current_app.config['RESULTS_FOLDER'],
                f"extraction_{job_id}.xlsx"
            )

            try:
                # Run extraction
                process_contracts(job_folder, output_file)

                # Count actual extracted results
                from openpyxl import load_workbook
                if os.path.exists(output_file):
                    wb = load_workbook(output_file)
                    contracts_count = max(0, wb['Contract Headers'].max_row - 1)
                    services_count = max(0, wb['Services & Rates'].max_row - 1)
                    job.total_contracts_extracted = contracts_count
                    job.total_services_extracted = services_count

                job.status = JobStatus.COMPLETED.value
                job.completed_at = datetime.utcnow()
                job.output_file_path = output_file
                job.processing_time_seconds = (
                    job.completed_at - job.started_at
                ).total_seconds()

            except Exception as e:
                job.status = JobStatus.FAILED.value
                job.error_message = str(e)
                logger.error(f"Extraction error: {e}")

            db.session.commit()

            # Log audit
            log_audit(user_id, 'EXTRACT', 'job', job_id, {'status': job.status})

            return jsonify({'message': 'Extraction complete', 'job': job.to_dict()}), 200

        except Exception as e:
            logger.error(f"Extract error: {e}")
            return jsonify({'error': 'Extraction failed'}), 500

    @app.route('/api/jobs')
    def get_jobs():
        """Get user's jobs"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']
            page = request.args.get('page', 1, type=int)
            per_page = current_app.config.get('ITEMS_PER_PAGE', 20)

            jobs = ExtractionJob.query.filter_by(user_id=user_id).paginate(
                page=page, per_page=per_page
            )

            return jsonify({
                'total': jobs.total,
                'pages': jobs.pages,
                'current_page': page,
                'jobs': [job.to_summary() for job in jobs.items]
            }), 200

        except Exception as e:
            logger.error(f"Get jobs error: {e}")
            return jsonify({'error': 'Failed to get jobs'}), 500

    @app.route('/api/jobs/<job_id>')
    def get_job(job_id):
        """Get job details"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']
            job = ExtractionJob.query.get(job_id)

            if not job:
                return jsonify({'error': 'Job not found'}), 404

            if job.user_id != user_id:
                return jsonify({'error': 'Unauthorized'}), 403

            return jsonify(job.to_dict()), 200

        except Exception as e:
            logger.error(f"Get job error: {e}")
            return jsonify({'error': 'Failed to get job'}), 500

    @app.route('/api/jobs/<job_id>/download')
    def download_results(job_id):
        """Download extraction results"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']
            job = ExtractionJob.query.get(job_id)

            if not job:
                return jsonify({'error': 'Job not found'}), 404

            if job.user_id != user_id:
                return jsonify({'error': 'Unauthorized'}), 403

            if not job.output_file_path or not os.path.exists(job.output_file_path):
                return jsonify({'error': 'Results file not found'}), 404

            # Log audit
            log_audit(user_id, 'DOWNLOAD', 'job', job_id)

            return send_file(
                job.output_file_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"extraction_{job_id}.xlsx"
            )

        except Exception as e:
            logger.error(f"Download error: {e}")
            return jsonify({'error': 'Failed to download'}), 500

    @app.route('/api/jobs', methods=['DELETE'])
    def delete_jobs():
        """Delete one or multiple jobs"""
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        try:
            user_id = session['user_id']
            data = request.get_json()
            job_ids = data.get('job_ids', [])

            if not job_ids or not isinstance(job_ids, list):
                return jsonify({'error': 'Invalid job_ids'}), 400

            deleted_count = 0
            failed = []

            for job_id in job_ids:
                job = ExtractionJob.query.get(job_id)

                if not job:
                    failed.append({'job_id': job_id, 'error': 'Not found'})
                    continue

                if job.user_id != user_id:
                    failed.append({'job_id': job_id, 'error': 'Unauthorized'})
                    continue

                try:
                    # Delete output file if exists
                    if job.output_file_path and os.path.exists(job.output_file_path):
                        os.remove(job.output_file_path)
                        logger.info(f"Deleted file: {job.output_file_path}")

                    # Delete ContractResult records (will cascade)
                    ContractResult.query.filter_by(job_id=job_id).delete()

                    # Delete the job
                    db.session.delete(job)

                    # Delete audit logs for this job
                    AuditLog.query.filter_by(resource_type='job', resource_id=job_id).delete()

                    # Log the deletion
                    log_audit(user_id, 'DELETE', 'job', job_id)

                    deleted_count += 1
                    logger.info(f"Deleted job: {job_id}")

                except Exception as e:
                    failed.append({'job_id': job_id, 'error': str(e)})
                    logger.error(f"Error deleting job {job_id}: {e}")

            db.session.commit()

            return jsonify({
                'deleted_count': deleted_count,
                'failed': failed,
                'message': f'Deleted {deleted_count} job(s)'
            }), 200

        except Exception as e:
            logger.error(f"Delete jobs error: {e}")
            db.session.rollback()
            return jsonify({'error': 'Failed to delete jobs'}), 500

    # ==================== Error Handlers ====================

    @app.errorhandler(404)
    def not_found(error):
        return render_template('404.html'), 404

    @app.errorhandler(500)
    def internal_error(error):
        logger.error(f"Internal error: {error}")
        return render_template('500.html'), 500

    return app


if __name__ == '__main__':
    app = create_app(os.getenv('FLASK_ENV', 'development'))
    app.run(
        host=os.getenv('FLASK_HOST', '0.0.0.0'),
        port=int(os.getenv('FLASK_PORT', 5000)),
        debug=os.getenv('FLASK_DEBUG', True)
    )
